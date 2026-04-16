from __future__ import annotations

from copy import copy
from typing import Optional
from openpyxl import load_workbook

HEADER_ROW = 1
DATA_START_ROW = 4
CRS620MI_TAB_PREFIX = "API_CRS620MI_"


def norm(v):
    if v is None:
        return ""
    return str(v).strip()


def build_maps(lookup_path: str):
    wb = load_workbook(lookup_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = {norm(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    suno_col = headers["SUNO"]
    newsuno_col = headers["NEWSUNO"]

    old_to_news = {}
    news_to_old = {}
    multi_old_keys = set()

    for r in range(2, ws.max_row + 1):
        old = norm(ws.cell(r, suno_col).value)
        new = norm(ws.cell(r, newsuno_col).value)
        if not old or not new:
            continue

        bucket = old_to_news.setdefault(old, [])
        if new not in bucket:
            bucket.append(new)
            if len(bucket) > 1:
                multi_old_keys.add(old)

        news_to_old[new] = old

    return old_to_news, news_to_old, multi_old_keys


def clone_row_styles(ws, src_row: int, dst_row: int, max_col: int):
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
    if ws.row_dimensions[src_row].height is not None:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def clear_row(ws, row_idx: int, max_col: int):
    for c in range(1, max_col + 1):
        ws.cell(row_idx, c).value = None


def _iter_non_empty_data_rows(ws, max_col: int):
    for row_values in ws.iter_rows(
        min_row=DATA_START_ROW,
        max_row=ws.max_row,
        min_col=1,
        max_col=max_col,
        values_only=True,
    ):
        if any(v is not None for v in row_values):
            yield list(row_values)


def process_sheet(ws, old_to_news, news_to_old, multi_old_keys):
    title = ws.title
    max_col = ws.max_column
    headers = {norm(ws.cell(HEADER_ROW, c).value): c for c in range(1, max_col + 1)}

    # Fast skip for irrelevant tabs.
    if title not in ("API_CRS620MI_UpdSupplier", "API_CRS620MI_CopyTemplate") and "SUNO" not in headers:
        return 0, 0

    suno_idx = headers.get("SUNO", 0) - 1
    suno_hash_idx = headers.get("SUNO#", 0) - 1
    cfi1_idx = headers.get("CFI1", 0) - 1

    data_rows = list(_iter_non_empty_data_rows(ws, max_col))
    if not data_rows:
        return 0, 0

    new_rows = []
    changed = False

    for values in data_rows:
        if title == "API_CRS620MI_UpdSupplier":
            if cfi1_idx < 0 or suno_idx < 0:
                new_rows.append(values)
                continue

            old = norm(values[cfi1_idx])
            if old not in multi_old_keys:
                new_rows.append(values)
                continue

            for new_suno in old_to_news.get(old, []):
                row_copy = values.copy()
                row_copy[suno_idx] = new_suno
                new_rows.append(row_copy)
            changed = True
            continue

        if suno_idx < 0:
            new_rows.append(values)
            continue

        current_suno = norm(values[suno_idx])
        old = news_to_old.get(current_suno, current_suno)

        if old not in multi_old_keys:
            new_rows.append(values)
            continue

        for new_suno in old_to_news.get(old, []):
            row_copy = values.copy()
            row_copy[suno_idx] = new_suno

            if title == "API_CRS620MI_CopyTemplate" and suno_hash_idx >= 0:
                row_copy[suno_hash_idx] = old
            elif suno_hash_idx >= 0 and not norm(values[suno_hash_idx]):
                row_copy[suno_hash_idx] = old

            new_rows.append(row_copy)

        changed = True

    if not changed and len(new_rows) == len(data_rows):
        return len(data_rows), len(new_rows)

    original_last = ws.max_row
    template_row = DATA_START_ROW if DATA_START_ROW <= ws.max_row else HEADER_ROW

    for i, row_values in enumerate(new_rows, start=DATA_START_ROW):
        # Existing rows already have style. Clone only for newly created rows.
        if i > original_last:
            clone_row_styles(ws, template_row, i, max_col)
        for c, value in enumerate(row_values, start=1):
            ws.cell(i, c).value = value

    for r in range(DATA_START_ROW + len(new_rows), original_last + 1):
        clear_row(ws, r, max_col)

    return len(data_rows), len(new_rows)


def expand_crs620mi_suno(target_path: str, lookup_path: str, output_path: Optional[str] = None):
    resolved_output = output_path or target_path
    old_to_news, news_to_old, multi_old_keys = build_maps(lookup_path)

    # If there are no one-to-many mappings, nothing to expand.
    if not multi_old_keys:
        return []

    wb = load_workbook(target_path)

    summary = []
    for ws in wb.worksheets:
        if ws.title == "Sheet1" or not ws.title.startswith(CRS620MI_TAB_PREFIX):
            continue
        before, after = process_sheet(ws, old_to_news, news_to_old, multi_old_keys)
        summary.append((ws.title, before, after))

    wb.save(resolved_output)
    return summary
