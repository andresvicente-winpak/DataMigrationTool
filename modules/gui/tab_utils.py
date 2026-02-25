import customtkinter as ctk
from tkinter import filedialog, messagebox, simpledialog
import threading
import openpyxl
import os
from colorama import Fore, Style
from modules.sdt_utils import SDTUtils
from modules.script_runner import ScriptRunner
from modules.gui.utils import bind_context_help # <--- IMPORT

class UtilitiesHub(ctk.CTkTabview):
    def __init__(self, parent):
        super().__init__(parent)
        self.add("Copy Sheet")
        self.add("Merge Files")
        self.add("Script Library") 
        
        self.script_runner = ScriptRunner()

        # --- BIND HELP TO TABS ---
        try:
            btns = self._segmented_button._buttons_dict
            if "Copy Sheet" in btns: bind_context_help(btns["Copy Sheet"], "Copy Tool: Move data between sheets with header auto-matching.")
            if "Merge Files" in btns: bind_context_help(btns["Merge Files"], "Merge Tool: Combine multiple Excel files into one Master file (deduplicating rows).")
            if "Script Library" in btns: bind_context_help(btns["Script Library"], "Script Launcher: Run standalone python scripts from the 'utilities/' folder.")
        except: pass

        self._build_copy(self.tab("Copy Sheet"))
        self._build_merge(self.tab("Merge Files"))
        self._build_scripts(self.tab("Script Library"))

    # --- COPY SHEET TAB ---
    def _build_copy(self, frame):
        ctk.CTkLabel(frame, text="Copy SDT Sheet Data", font=("Arial", 18, "bold")).pack(anchor="w", pady=10)
        
        # File Selection
        self.file_entry = ctk.CTkEntry(frame, placeholder_text="Select SDT File")
        self.file_entry.pack(fill="x", pady=5)
        
        btn_load = ctk.CTkButton(frame, text="Browse & Load", command=self._load_file)
        btn_load.pack(anchor="e", pady=5)
        bind_context_help(btn_load, "Open an Excel file and read its sheet names.")
        
        # Dropdowns
        ctk.CTkLabel(frame, text="Source Sheet:").pack(anchor="w", pady=(10,0))
        self.src_var = ctk.StringVar(value="")
        self.src_menu = ctk.CTkOptionMenu(frame, variable=self.src_var, values=["Load File First"])
        self.src_menu.pack(fill="x")
        bind_context_help(self.src_menu, "The sheet containing data you want to copy FROM.")
        
        ctk.CTkLabel(frame, text="Destination Sheet:").pack(anchor="w", pady=(10,0))
        self.dst_var = ctk.StringVar(value="")
        self.dst_menu = ctk.CTkOptionMenu(frame, variable=self.dst_var, values=["Load File First"])
        self.dst_menu.pack(fill="x")
        bind_context_help(self.dst_menu, "The sheet you want to copy TO.")
        
        # Option
        self.nok_var = ctk.BooleanVar(value=True)
        chk = ctk.CTkCheckBox(frame, text="Include 'NOK' rows", variable=self.nok_var)
        chk.pack(anchor="w", pady=15)
        bind_context_help(chk, "If checked, rows marked as 'NOK' (Failed) in the Message column will be copied.\nUncheck to copy only successful/new rows.")
        
        btn_copy = ctk.CTkButton(frame, text="COPY DATA", fg_color="orange", text_color="black", height=40, command=self._run_copy)
        btn_copy.pack(fill="x", pady=10)
        bind_context_help(btn_copy, "Execute copy. Matches columns by name and appends rows to Destination.")
        
        self.wb = None
        self.filepath = None

    def _load_file(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not f: return
        self.file_entry.delete(0,"end"); self.file_entry.insert(0, f)
        self.filepath = f
        
        try:
            self.wb = openpyxl.load_workbook(f)
            sheets = self.wb.sheetnames
            self.src_menu.configure(values=sheets); self.src_var.set(sheets[0])
            self.dst_menu.configure(values=sheets); self.dst_var.set(sheets[0])
            print(f"Loaded {os.path.basename(f)}")
        except Exception as e:
            print(f"Error loading file: {e}")

    def _run_copy(self):
        if not self.wb or not self.filepath: return
        
        src = self.src_var.get()
        dst = self.dst_var.get()
        include_nok = self.nok_var.get() # Get checkbox value
        
        def _t():
            try:
                utils = SDTUtils()
                ws_src = self.wb[src]
                ws_dst = self.wb[dst]
                
                print(f"Copying {src} -> {dst} (Include NOK: {include_nok})...")
                
                matched, copied = utils._map_and_copy_data(ws_src, ws_dst, include_nok=include_nok)
                
                self.wb.save(self.filepath)
                print(f"{Fore.GREEN}Success! Copied {copied} rows.{Style.RESET_ALL}")
                
            except Exception as e:
                print(f"{Fore.RED}Error: {e}{Style.RESET_ALL}")
                
        threading.Thread(target=_t).start()

    # --- MERGE FILES TAB ---
    def _build_merge(self, frame):
        ctk.CTkLabel(frame, text="Merge Files (Deduplicate)", font=("Arial", 18, "bold")).pack(anchor="w", pady=10)

        # Master File
        ctk.CTkLabel(frame, text="1. Master File (Destination):").pack(anchor="w")
        self.merge_master_ent = ctk.CTkEntry(frame)
        self.merge_master_ent.pack(fill="x", pady=5)
        
        btn_mast = ctk.CTkButton(frame, text="Select Master", command=self._sel_master)
        btn_mast.pack(anchor="e")
        bind_context_help(btn_mast, "Select the main file that aggregates all data.")

        # Source File
        ctk.CTkLabel(frame, text="2. Source File (Data to Add):").pack(anchor="w", pady=(10,0))
        self.merge_src_ent = ctk.CTkEntry(frame)
        self.merge_src_ent.pack(fill="x", pady=5)
        
        btn_src = ctk.CTkButton(frame, text="Select Source", command=self._sel_src)
        btn_src.pack(anchor="e")
        bind_context_help(btn_src, "Select the file containing new rows to merge.")

        # Sheet Select
        ctk.CTkLabel(frame, text="Select Common Sheet:").pack(anchor="w", pady=(10,0))
        self.merge_sheet_var = ctk.StringVar(value="Select Files First")
        self.merge_sheet_menu = ctk.CTkOptionMenu(frame, variable=self.merge_sheet_var)
        self.merge_sheet_menu.pack(fill="x", pady=5)
        bind_context_help(self.merge_sheet_menu, "Select the sheet name present in BOTH files to merge.")
        
        btn_merge = ctk.CTkButton(frame, text="MERGE & SAVE", fg_color="green", height=40, command=self._run_merge)
        btn_merge.pack(fill="x", pady=20)
        bind_context_help(btn_merge, "Reads source, checks for duplicates, and appends unique rows to Master.\nSaves the Master file.")
        
        self.wb_master = None
        self.path_master = None
        self.path_source = None

    def _sel_master(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if f:
            self.merge_master_ent.delete(0,"end"); self.merge_master_ent.insert(0,f)
            self.path_master = f
            try:
                self.wb_master = openpyxl.load_workbook(f)
                print(f"Master Loaded: {os.path.basename(f)}")
            except Exception as e: print(f"Error: {e}")

    def _sel_src(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if f:
            self.merge_src_ent.delete(0,"end"); self.merge_src_ent.insert(0,f)
            self.path_source = f
            self._update_common_sheets()

    def _update_common_sheets(self):
        if not self.wb_master or not self.path_source: return
        try:
            wb_src = openpyxl.load_workbook(self.path_source, read_only=True)
            src_sheets = wb_src.sheetnames
            master_sheets = self.wb_master.sheetnames
            
            common = [s for s in src_sheets if s in master_sheets]
            if common:
                self.merge_sheet_menu.configure(values=common)
                self.merge_sheet_var.set(common[0])
            else:
                self.merge_sheet_menu.configure(values=["No Common Sheets"])
                self.merge_sheet_var.set("No Common Sheets")
        except Exception as e: print(f"Error reading source: {e}")

    def _run_merge(self):
        if not self.wb_master or not self.path_source: return
        
        sheet = self.merge_sheet_var.get()
        if sheet == "No Common Sheets": return

        def _t():
            try:
                utils = SDTUtils()
                wb_src = openpyxl.load_workbook(self.path_source, data_only=True)
                
                print(f"Merging '{sheet}'...")
                count = utils._merge_sheet_data(self.wb_master[sheet], wb_src[sheet])
                
                print(f"Saving Master...")
                self.wb_master.save(self.path_master)
                print(f"{Fore.GREEN}Success! Added {count} unique rows.{Style.RESET_ALL}")
            except Exception as e:
                # This print now works because Fore/Style are imported
                print(f"Merge Error: {e}")
            
        threading.Thread(target=_t).start()

    # --- NEW: SCRIPT LIBRARY TAB ---
    def _build_scripts(self, frame):
        ctk.CTkLabel(frame, text="External Script Library", font=("Arial", 18, "bold")).pack(anchor="w", pady=10)
        ctk.CTkLabel(frame, text="Run standalone CLI utilities from the 'utilities/' folder.", text_color="gray").pack(anchor="w", pady=(0, 10))

        # Refresh Button
        btn_ref = ctk.CTkButton(frame, text="Refresh Library", width=100, command=self._refresh_scripts)
        btn_ref.pack(anchor="e", padx=10)
        bind_context_help(btn_ref, "Re-scan the 'utilities/' folder for new scripts.")

        # Scrollable Area for Categories
        self.script_scroll = ctk.CTkScrollableFrame(frame)
        self.script_scroll.pack(fill="both", expand=True, pady=10)
        
        self._refresh_scripts()

    def _refresh_scripts(self):
        # Clear existing
        for widget in self.script_scroll.winfo_children(): widget.destroy()
        
        data = self.script_runner.scan_scripts()
        
        if not data:
            ctk.CTkLabel(self.script_scroll, text="No scripts found in 'utilities/' folder.").pack(pady=20)
            ctk.CTkLabel(self.script_scroll, text="Create subfolders (e.g., 'utilities/Items') and drop .py files there.").pack()
            return

        for category, scripts in data.items():
            # Category Header
            cat_frame = ctk.CTkFrame(self.script_scroll, fg_color="transparent")
            cat_frame.pack(fill="x", pady=(10, 5))
            ctk.CTkLabel(cat_frame, text=f"📂 {category}", font=("Arial", 14, "bold"), text_color="#3B8ED0").pack(anchor="w")
            
            # Script Rows
            for script in scripts:
                row = ctk.CTkFrame(self.script_scroll, fg_color="#2B2B2B")
                row.pack(fill="x", pady=2, padx=10)
                
                # Script Name & Desc
                info_frame = ctk.CTkFrame(row, fg_color="transparent")
                info_frame.pack(side="left", fill="x", expand=True, padx=10, pady=5)
                
                ctk.CTkLabel(info_frame, text=script['name'], font=("Consolas", 12, "bold")).pack(anchor="w")
                
                # Description logic (clickable to edit)
                desc_txt = script['description']
                if desc_txt == "No description provided.": desc_col = "gray"
                else: desc_col = "#AAAAAA"
                
                desc_lbl = ctk.CTkButton(info_frame, text=desc_txt, fg_color="transparent", 
                                         text_color=desc_col, anchor="w", hover_color="#333333",
                                         command=lambda s=script: self._edit_desc(s))
                desc_lbl.pack(fill="x")
                bind_context_help(desc_lbl, "Click to edit this script's description.")

                # Run Button
                btn_run = ctk.CTkButton(row, text="▶ RUN CLI", width=80, fg_color="green", 
                              command=lambda p=script['path']: self._run_script(p))
                btn_run.pack(side="right", padx=10)
                bind_context_help(btn_run, "Launch this script in a new terminal window.")

    def _edit_desc(self, script):
        new_desc = simpledialog.askstring("Edit Description", f"Description for {script['name']}:", initialvalue=script['description'])
        if new_desc:
            self.script_runner.update_description(script['id'], new_desc)
            self._refresh_scripts()

    def _run_script(self, path):
        err = self.script_runner.run_script_in_terminal(path)
        if err:
            messagebox.showerror("Error", f"Failed to launch script: {err}")