import openpyxl
from colorama import Fore, Style

class SDTUtils:
    def __init__(self):
        pass

    def _map_and_copy_data(self, source_ws, dest_ws, include_nok=True):
        """
        Copies data from source_ws to dest_ws based on matching column headers.
        """
        # 1. Map Headers (Row 1)
        src_headers = {}
        for col_idx, cell in enumerate(source_ws[1], start=1):
            if cell.value:
                src_headers[str(cell.value).strip().upper()] = col_idx

        dst_map = {} # Map Src_Col_Idx -> Dst_Col_Idx
        for col_idx, cell in enumerate(dest_ws[1], start=1):
            if cell.value:
                header = str(cell.value).strip().upper()
                if header in src_headers:
                    dst_map[src_headers[header]] = col_idx

        if not dst_map:
            print("No matching columns found.")
            return 0, 0

        msg_col_idx = src_headers.get('MESSAGE')

        # 2. Copy Data (Start Row 2)
        copied_count = 0
        
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            # Check Filtering
            if not include_nok and msg_col_idx:
                msg_val = str(row[msg_col_idx - 1]).upper() if len(row) >= msg_col_idx else ""
                if msg_val.startswith('NOK'):
                    continue

            # Create new row data
            max_dst_col = dest_ws.max_column
            # Initialize with None
            new_row_data = [None] * max_dst_col
            
            has_data = False
            for src_idx, dst_idx in dst_map.items():
                val = row[src_idx - 1]
                if dst_idx - 1 < len(new_row_data):
                    new_row_data[dst_idx - 1] = val
                    has_data = True
            
            if has_data:
                dest_ws.append(new_row_data)
                copied_count += 1

        return len(dst_map), copied_count

    def _merge_sheet_data(self, master_ws, source_ws):
        master_hashes = set()
        for row in master_ws.iter_rows(min_row=2, values_only=True):
            row_hash = hash(tuple(str(x) for x in row if x is not None))
            master_hashes.add(row_hash)

        added = 0
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            row_hash = hash(tuple(str(x) for x in row if x is not None))
            if row_hash not in master_hashes:
                master_ws.append(list(row))
                master_hashes.add(row_hash)
                added += 1
                
        return added