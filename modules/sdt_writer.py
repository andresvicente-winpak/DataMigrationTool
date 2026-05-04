import os
import shutil
import pandas as pd
from colorama import Fore, Style
from openpyxl import load_workbook
from modules.transform_engine import TransformEngine

class SDTWriter:
    def __init__(self, output_dir='output'):
        self.output_dir = output_dir
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def _transform_data(self, df_legacy, rules_df, valid_fields, sheet_name):
        """
        Applies rules to legacy data to create the target DataFrame.
        """
        # Ensure TARGET_FIELD exists to prevent KeyError
        if 'TARGET_FIELD' not in rules_df.columns:
            return pd.DataFrame(index=df_legacy.index)

        relevant_rules = rules_df[rules_df['TARGET_FIELD'].isin(valid_fields)].copy()

        if relevant_rules.empty:
            return pd.DataFrame(index=df_legacy.index)

        engine = TransformEngine(relevant_rules, lookups={})
        df_transformed = engine.process(df_legacy)

        return df_transformed

    @staticmethod
    def _norm_cell(x):
        """
        Normalizes cell values before writing to Excel:
        - NaN -> ""
        - "0.0" -> "0" (and generally any numeric ending with .0)
        - trims whitespace
        - escapes leading formula characters so text is stored as literal text
        """
        if pd.isna(x):
            return ""

        s = str(x).strip()

        # Convert "0.0" / "1.0" / "123.0" -> "0" / "1" / "123"
        if s.endswith(".0"):
            try:
                return str(int(float(s)))
                s = str(int(float(s)))
            except Exception:
                return s
                pass

        # Prevent accidental formula injection / invalid formulas in output.
        # openpyxl treats values beginning with '=' as formulas.
        # We force these to literal strings so Excel won't repair the workbook.
        if s.startswith(('=', '+', '-', '@')):
            return f"'{s}"

        return s
        
    def _validate_written_rows(self, ws, sheet_name, start_row, valid_fields, expected_rows):
        """
        Validates that written worksheet values match the transformed rule output values.
        """
        mismatches = []

        for row_offset, expected_row in enumerate(expected_rows):
            excel_row = start_row + row_offset
            for col_offset, expected_value in enumerate(expected_row, start=1):
                actual_value = ws.cell(row=excel_row, column=col_offset).value
                actual_str = "" if actual_value is None else str(actual_value).strip()
                expected_str = "" if expected_value is None else str(expected_value).strip()
                if actual_str != expected_str:
                    mismatches.append({
                        "row": excel_row,
                        "column": valid_fields[col_offset - 1],
                        "expected": expected_str,
                        "actual": actual_str,
                    })

        if mismatches:
            first = mismatches[0]
            print(
                f"{Fore.YELLOW}   [Rule Check] {sheet_name}: {len(mismatches)} cell mismatch(es). "
                f"First mismatch row {first['row']} col {first['column']} (expected='{first['expected']}' actual='{first['actual']}').{Style.RESET_ALL}"
            )
            status = "warning"
        else:
            print(f"{Fore.GREEN}   [Rule Check] {sheet_name}: All generated cell values match rule output.{Style.RESET_ALL}")
            status = "ok"

        return {
            "sheet": sheet_name,
            "status": status,
            "rows_checked": len(expected_rows),
            "cells_checked": len(expected_rows) * len(valid_fields),
            "mismatch_count": len(mismatches),
            "mismatches": mismatches[:10],
        }
        
    def generate_from_template(
        self,
        template_path,
        legacy_data,
        rules_df,
        target_sheets,
        output_filename,
        append_if_exists=False
    ):
        """
        Creates a new SDT file based on the template and populated with data.
        If append_if_exists=True and file exists, it appends to the existing file instead of overwriting.
        """
        output_path = os.path.join(self.output_dir, output_filename)

        try:
            # 1. Handle File Creation vs Appending
            if append_if_exists and os.path.exists(output_path):
                print(f"   -> Appending to existing file: {output_filename}")
            else:
                # Create fresh copy
                shutil.copy(template_path, output_path)

            # 2. Open Workbook
            wb = load_workbook(output_path)
            
            validation_summary = []

            for sheet_name in target_sheets:
                if sheet_name not in wb.sheetnames:
                    print(f"{Fore.YELLOW}   [Warning] Sheet '{sheet_name}' not found in template.{Style.RESET_ALL}")
                    continue

                ws = wb[sheet_name]

                # 3. Read Header from Template (Row 1 assumed)
                headers = [cell.value for cell in ws[1]]
                valid_fields = [str(h).strip() for h in headers if h]

                # 4. Transform Data
                df_out = self._transform_data(legacy_data, rules_df, valid_fields, sheet_name)

                # 5. Align Columns
                df_final = pd.DataFrame(index=df_out.index)

                for field in valid_fields:
                    if field in df_out.columns:
                        df_final[field] = df_out[field]
                    else:
                        df_final = df_out.reindex(columns=valid_fields, fill_value="")

                # 6. Write to Sheet (Append to end)
                data_rows = df_final[valid_fields].values.tolist()
                start_row = ws.max_row + 1
                normalized_rows = []

                for row_data in data_rows:
                    clean_row = [self._norm_cell(x) for x in row_data]
                    ws.append(clean_row)
                    normalized_rows.append(clean_row)

                validation_summary.append(
                    self._validate_written_rows(ws, sheet_name, start_row, valid_fields, normalized_rows)
                )
                
            wb.save(output_path)

            if not append_if_exists:
                print(f"   -> Generated: {output_path}")
                
            return {"output_path": output_path, "sheets": validation_summary}

        except Exception as e:
            print(f"{Fore.RED}   [SDT WRITER ERROR] {e}{Style.RESET_ALL}")
            raise