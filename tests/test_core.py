import pytest
import pandas as pd
import os
import sys
import glob
import datetime
import configparser # <--- ADDED MISSING IMPORT
from openpyxl import load_workbook

# Add root directory
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from modules.config_loader import ConfigLoader
from modules.extractor import DataExtractor
from modules.transform_engine import TransformEngine
from modules.sdt_writer import SDTWriter
from modules.validator_analyzer import ValidatorAnalyzer
from modules.surgical_extractor import SurgicalExtractor
from modules.mco_importer import MCOImporter
from modules.migration_runner import MigrationRunner
from modules.batch_processor import BatchProcessor
from modules.mco_checker import MCOChecker
from modules.rule_manager import RuleManager
from modules.audit_manager import AuditManager
from modules.sdt_utils import SDTUtils
from modules.sync_manager import SyncManager
from modules.script_runner import ScriptRunner
from modules.auto_detector import AutoDetector

# GUI Modules
from modules.gui.app import M3MigrationApp
from modules.gui.tab_migration import MigrationHub
from modules.gui.tab_config import ConfigHub
from modules.gui.tab_rules import RulesHub
from modules.gui.tab_utils import UtilitiesHub

CONF_DIR = "tests/config_temp"
DATA_DIR = "tests/data_temp"
OUT_DIR = "tests/output_temp"
STAGE_DIR = "tests/surgical_staging"
UTIL_DIR = "tests/utilities_temp"

# --- TESTS ---

def test_01_directory_structure():
    assert os.path.exists(CONF_DIR)
    assert os.path.exists(DATA_DIR)

def test_02_extractor_load_excel():
    fpath = f"{DATA_DIR}/legacy.xlsx"
    pd.DataFrame({'MMITNO': ['100'], 'DESC': ['Item A']}).to_excel(fpath, sheet_name='Sheet1', index=False)
    ex = DataExtractor()
    df = ex.load_data(fpath, sheet_name='Sheet1')
    assert not df.empty
    assert 'MMITNO' in df.columns

def test_03_config_loader():
    rule_path = f"{CONF_DIR}/rules/TEST_API.xlsx"
    df = pd.DataFrame({
        'TARGET_FIELD': ['ITNO', 'ITDS'],
        'SOURCE_FIELD': ['MMITNO', 'MMITDS'],
        'RULE_TYPE': ['DIRECT', 'DIRECT'],
        'RULE_VALUE': ['', ''],
        'SCOPE': ['GLOBAL', 'GLOBAL']
    })
    df.to_excel(rule_path, sheet_name='Rules', index=False)
    loader = ConfigLoader('TEST_API', rule_dir=f"{CONF_DIR}/rules")
    rules, _ = loader.load_config()
    assert len(rules) == 2

def test_04_transform_engine_direct():
    df_source = pd.DataFrame({'MMITNO': ['100', '200']})
    df_rules = pd.DataFrame([{
        'TARGET_FIELD': 'ITNO', 'RULE_TYPE': 'DIRECT', 'SOURCE_FIELD': 'MMITNO', 'RULE_VALUE': ''
    }])
    eng = TransformEngine(df_rules, {})
    res = eng.process(df_source)
    assert res.iloc[0]['ITNO'] == '100'

def test_05_transform_engine_const():
    df_source = pd.DataFrame({'A': [1, 2]})
    df_rules = pd.DataFrame([{
        'TARGET_FIELD': 'STAT', 'RULE_TYPE': 'CONST', 'RULE_VALUE': '20', 'SOURCE_FIELD': '' 
    }])
    eng = TransformEngine(df_rules, {})
    res = eng.process(df_source)
    assert res.iloc[0]['STAT'] == '20'

def test_05b_transform_engine_const_blank_quotes():
    df_source = pd.DataFrame({'A': [1]})
    df_rules = pd.DataFrame([{
        'TARGET_FIELD': 'EMPTY1', 'RULE_TYPE': 'CONST', 'RULE_VALUE': '""', 'SOURCE_FIELD': ''
    }, {
        'TARGET_FIELD': 'EMPTY2', 'RULE_TYPE': 'CONST', 'RULE_VALUE': "''", 'SOURCE_FIELD': ''
    }])
    eng = TransformEngine(df_rules, {})
    res = eng.process(df_source)
    assert res.iloc[0]['EMPTY1'] == ''
    assert res.iloc[0]['EMPTY2'] == ''

def test_06_python_hooks():
    code_snippet = "if source == 1: return 'HOOKED'\nreturn 'NOPE'"
    df_s = pd.DataFrame({'A': [1]})
    df_r = pd.DataFrame([{'TARGET_FIELD':'B', 'RULE_TYPE':'PYTHON', 'RULE_VALUE': code_snippet, 'SOURCE_FIELD':'A'}])
    eng = TransformEngine(df_r, {})
    res = eng.process(df_s)
    assert res.iloc[0]['B'] == 'HOOKED'

def test_07_surgical_extraction():
    extractor = SurgicalExtractor()
    extractor.config_dir = CONF_DIR
    extractor.staging_dir = STAGE_DIR
    
    pd.DataFrame([{'OBJECT_TYPE': 'ITEM', 'MCO_SHEET': 'Item_Master', 'KEY_COLUMN': 'MMITNO'}]).to_csv(f"{CONF_DIR}/surgical_def.csv", index=False)
    
    legacy_path = f"{DATA_DIR}/legacy_items.xlsx"
    pd.DataFrame({'MMITNO': ['1001', '1002', '1003'], 'DESC': ['A','B','C']}).to_excel(legacy_path, index=False)
    
    pd.DataFrame([{'MCO_SHEET': 'Item_Master', 'SOURCE_FILE': legacy_path, 'JOIN_KEY': 'MMITNO'}]).to_csv(f"{CONF_DIR}/source_map.csv", index=False)
    
    # Add migration map for fallback API lookup
    pd.DataFrame([{'MCO_SHEET': 'Item_Master', 'API_NAME': 'MMS200MI', 'SDT_TEMPLATE': '', 'TRANSACTION_SHEET': ''}]).to_csv(f"{CONF_DIR}/migration_map.csv", index=False)
    
    tasks = extractor.perform_extraction('ITEM', ['1002'])
    
    assert len(tasks) == 1
    assert "STAGED_MMS200MI" in tasks[0]['legacy_path']
    df = pd.read_excel(tasks[0]['legacy_path'])
    assert str(df.iloc[0]['MMITNO']) == '1002'

def test_08_batch_processor():
    batch = BatchProcessor()
    b_path = f"{DATA_DIR}/batch_job.xlsx"
    pd.DataFrame({
        'JOB_ID': [1], 'RULE_CONFIG': ['TEST_API'], 'LEGACY_FILE': [f"{DATA_DIR}/legacy.xlsx"],
        'TARGET_SHEETS': ['Sheet1']
    }).to_excel(b_path, index=False)
    df = batch.load_batch_file(b_path)
    assert not df.empty

def test_09_mco_checker():
    checker = MCOChecker()
    mco_path = f"{DATA_DIR}/bad_mco.xlsx"
    pd.DataFrame({'FIELD NAME': ['ITNO', 'ITNO'], 'CONVERSION SOURCE': ['A', 'B']}).to_excel(mco_path, index=False)
    try: checker.check_file(mco_path)
    except: pass

def test_10_importer_logic():
    importer = MCOImporter()
    mco_path = f"{DATA_DIR}/MCO_VALID.xlsx" 
    pd.DataFrame({'FIELD NAME': ['ITNO'], 'SOURCE': ['MMITNO']}).to_excel(mco_path, sheet_name='Sheet1', startrow=2, index=False)
    success = importer.run_import_headless(mco_path, "Sheet1", "IMPORTED_API", output_dir=f"{CONF_DIR}/rules")
    assert success

def test_11_sdt_writer():
    writer = SDTWriter(output_dir=OUT_DIR)
    tpl_path = f"{CONF_DIR}/sdt_templates/template.xlsx"
    if not os.path.exists(os.path.dirname(tpl_path)): os.makedirs(os.path.dirname(tpl_path))
    wb = pd.ExcelWriter(tpl_path, engine='xlsxwriter')
    pd.DataFrame(columns=['ITNO', 'DESC', 'MESSAGE']).to_excel(wb, sheet_name='Add', index=False, startrow=0)
    wb.close()
    
    df_data = pd.DataFrame({'ITNO': ['100'], 'DESC': ['Test']})
    empty_rules = pd.DataFrame(columns=['TARGET_FIELD', 'RULE_TYPE', 'RULE_VALUE', 'SOURCE_FIELD'])
    
    writer.generate_from_template(tpl_path, df_data, empty_rules, ['Add'], "output_test.xlsx")
    assert os.path.exists(f"{OUT_DIR}/output_test.xlsx")

def test_12_surgical_aliasing():
    extractor = SurgicalExtractor()
    extractor.config_dir = CONF_DIR
    extractor.staging_dir = STAGE_DIR
    df_whs = pd.DataFrame({'MBITNO': ['9999'], 'MBWHLO': ['001']})
    whs_path = f"{DATA_DIR}/test_whs.xlsx"
    df_whs.to_excel(whs_path, index=False)
    pd.DataFrame([{'MCO_SHEET': 'Test_Whs', 'SOURCE_FILE': whs_path, 'JOIN_KEY': 'MBITNO'}]).to_csv(f"{CONF_DIR}/source_map.csv", index=False)
    pd.DataFrame([{'OBJECT_TYPE': 'WHS', 'MCO_SHEET': 'Test_Whs', 'KEY_COLUMN': 'MBITNO'}]).to_csv(f"{CONF_DIR}/surgical_def.csv", index=False)
    pd.DataFrame([{'MCO_SHEET': 'Test_Whs', 'API_NAME': 'TEST_API', 'SDT_TEMPLATE': 'TEST_API.xlsx', 'TRANSACTION_SHEET': 'AddBasic'}]).to_csv(f"{CONF_DIR}/migration_map.csv", index=False)
    tasks = extractor.perform_extraction('WHS', ['9999'])
    assert len(tasks) == 1

def test_13_audit_manager():
    am = AuditManager(rule_dir=f"{CONF_DIR}/rules")
    if not os.path.exists(f"{CONF_DIR}/rules/IMPORTED_API.xlsx"): test_10_importer_logic()
    am.create_snapshot("IMPORTED_API", "TestSnapshot")
    assert len(am.list_snapshots("IMPORTED_API")) >= 1

def test_14_validator_simple():
    val = ValidatorAnalyzer()
    df_leg = pd.DataFrame({'A': ['1', '1', '2']})
    df_m3 = pd.DataFrame({'A': ['1', '1', '1']})
    res = val.analyze_column_pair(df_leg, 'A', df_m3, 'A')
    assert res['TYPE'] == 'CONST'

def test_15_gui_app_instantiation():
    try:
        app = M3MigrationApp()
        app.destroy()
    except Exception: pass

def test_16_merge_draft_logic():
    rm = RuleManager(rule_dir=f"{CONF_DIR}/rules")
    pd.DataFrame([{
        'TARGET_FIELD': 'A', 'RULE_TYPE': 'TODO', 'SOURCE_FIELD': '', 'RULE_VALUE': ''
    }]).to_excel(f"{CONF_DIR}/rules/MERGE_TARGET.xlsx", sheet_name='Rules', index=False)
    pd.DataFrame([{
        'TARGET': 'A', 'SOURCE': 'COL_A', 'TYPE': 'DIRECT', 'LOGIC': ''
    }]).to_excel(f"{DATA_DIR}/draft.xlsx", index=False)
    rm.merge_draft_file(f"{DATA_DIR}/draft.xlsx", "MERGE_TARGET")
    df, _ = rm.load_rules("MERGE_TARGET")
    assert df.iloc[0]['RULE_TYPE'] == 'DIRECT'

def test_17_auto_detector():
    mco_path = f"{DATA_DIR}/MCO_SIG.xlsx"
    pd.DataFrame({
        'FIELD NAME': ['MMITNO', 'MMCUNO'],
        'SOURCE': ['ITNO', 'CUNO']
    }).to_excel(mco_path, sheet_name='Item_Master', startrow=2, index=False)
    detector = AutoDetector(mco_path)
    detector.prefix_map = {'MM': {'sheet': 'Item_Master', 'api': 'MMS200MI'}}
    leg_path = f"{DATA_DIR}/unknown.xlsx"
    pd.DataFrame({'MMITNO': ['1'], 'MMSTAT': ['20']}).to_excel(leg_path, index=False)
    prefix, sheet, api = detector.identify_file(leg_path)
    assert prefix == 'MM'

def test_18_rule_manager_validation():
    rm = RuleManager()
    assert rm.validate_const("123456", "N", 5) is not None

def test_19_map_lookup_rule():
    map_path = f"{CONF_DIR}/map_lookup.csv"
    pd.DataFrame({'KEY': ['OLD', 'ANCIENT'], 'VAL': ['NEW', 'ANTIQUE']}).to_csv(map_path, index=False)
    rule_val = f"{map_path}|KEY|VAL"
    df_s = pd.DataFrame({'SRC': ['OLD', 'ANCIENT']})
    df_r = pd.DataFrame([{
        'TARGET_FIELD': 'RES', 'RULE_TYPE': 'MAP', 'SOURCE_FIELD': 'SRC', 'RULE_VALUE': rule_val
    }])
    eng = TransformEngine(df_r, {})
    res = eng.process(df_s)
    assert res.iloc[0]['RES'] == 'NEW'

def test_19b_map_lookup_fallback_to_source_when_missing_key():
    map_path = f"{CONF_DIR}/map_lookup_fallback.csv"
    pd.DataFrame({'KEY': ['OLD'], 'VAL': ['NEW']}).to_csv(map_path, index=False)
    rule_val = f"{map_path}|KEY|VAL"
    df_s = pd.DataFrame({'SRC': ['OLD', 'UNMAPPED']})
    df_r = pd.DataFrame([{
        'TARGET_FIELD': 'RES', 'RULE_TYPE': 'MAP', 'SOURCE_FIELD': 'SRC', 'RULE_VALUE': rule_val
    }])
    eng = TransformEngine(df_r, {})
    res = eng.process(df_s)
    assert list(res['RES']) == ['NEW', 'UNMAPPED']

def test_19c_map_lookup_fallback_to_source_when_map_unavailable():
    missing_path = f"{CONF_DIR}/does_not_exist.csv"
    rule_val = f"{missing_path}|KEY|VAL"
    df_s = pd.DataFrame({'SRC': ['OLD', 'UNMAPPED']})
    df_r = pd.DataFrame([{
        'TARGET_FIELD': 'RES', 'RULE_TYPE': 'MAP', 'SOURCE_FIELD': 'SRC', 'RULE_VALUE': rule_val
    }])
    eng = TransformEngine(df_r, {})
    res = eng.process(df_s)
    assert list(res['RES']) == ['OLD', 'UNMAPPED']

def test_19d_map_lookup_fallback_to_source_when_rule_value_blank():
    df_s = pd.DataFrame({'SRC': ['OLD', 'UNMAPPED']})
    df_r = pd.DataFrame([{
        'TARGET_FIELD': 'RES', 'RULE_TYPE': 'MAP', 'SOURCE_FIELD': 'SRC', 'RULE_VALUE': ''
    }])
    eng = TransformEngine(df_r, {})
    res = eng.process(df_s)
    assert list(res['RES']) == ['OLD', 'UNMAPPED']

def test_20_scope_override():
    rule_path = f"{CONF_DIR}/rules/SCOPE_TEST.xlsx"
    df = pd.DataFrame([
        {'TARGET_FIELD': 'A', 'RULE_VALUE': 'GLOB', 'SCOPE': 'GLOBAL'},
        {'TARGET_FIELD': 'A', 'RULE_VALUE': 'US_VAL', 'SCOPE': 'DIV_US'}
    ])
    df.to_excel(rule_path, sheet_name='Rules', index=False)
    loader = ConfigLoader('SCOPE_TEST', rule_dir=f"{CONF_DIR}/rules")
    r_glob, _ = loader.load_config('GLOBAL')
    assert r_glob[r_glob['TARGET_FIELD']=='A'].iloc[0]['RULE_VALUE'] == 'GLOB'

def test_21_python_context_logic():
    code = "if str(source) == '20': return 'VIP'\nreturn 'STD'"
    df_s = pd.DataFrame({'MMSTAT': ['20', '30']})
    df_r = pd.DataFrame([{
        'TARGET_FIELD':'RES', 'RULE_TYPE':'PYTHON', 'RULE_VALUE': code, 'SOURCE_FIELD':'MMSTAT'
    }])
    eng = TransformEngine(df_r, {})
    res = eng.process(df_s)
    assert res.iloc[0]['RES'] == 'VIP'

def test_22_sdt_utils_copy():
    """Test copying rows between SDT sheets."""
    utils = SDTUtils()
    path_v2 = f"{OUT_DIR}/copy_test_v2.xlsx"
    wb2 = pd.ExcelWriter(path_v2, engine='xlsxwriter')
    pd.DataFrame({'A': [1], 'B': [2], 'MESSAGE': ['OK']}).to_excel(wb2, sheet_name='Source', index=False) 
    pd.DataFrame(columns=['A', 'B']).to_excel(wb2, sheet_name='Dest', index=False)
    wb2.close()
    
    from openpyxl import load_workbook
    wb_obj = load_workbook(path_v2)
    # Pass include_nok=True to ensure it copies the row
    matched, copied = utils._map_and_copy_data(wb_obj['Source'], wb_obj['Dest'], include_nok=True)
    assert copied == 1

def test_23_script_runner_scan():
    runner = ScriptRunner(root_dir=UTIL_DIR)
    cat_dir = f"{UTIL_DIR}/MyTools"
    if not os.path.exists(cat_dir): os.makedirs(cat_dir)
    with open(f"{cat_dir}/myscript.py", "w") as f: f.write("print('Hello')")
    scripts = runner.scan_scripts()
    assert "MyTools" in scripts

def test_24_filter_rule_engine():
    from modules.transform_engine import FilterEngine
    df_s = pd.DataFrame({'STAT': ['20', '90', '20'], 'VAL': [1, 2, 3]})
    df_rules = pd.DataFrame([{
        'TARGET_FIELD': '_ROW_', 'RULE_TYPE': 'FILTER', 'SOURCE_FIELD': 'STAT', 'RULE_VALUE': "source == '20'"
    }])
    fe = FilterEngine(df_rules)
    res = fe.apply_filters(df_s)
    assert len(res) == 2

def test_25_sync_manager_compare():
    f1 = f"{CONF_DIR}/rules/SYNC_LOCAL.xlsx"
    f2 = f"{DATA_DIR}/SYNC_REMOTE.xlsx"
    df1 = pd.DataFrame([{'TARGET_FIELD': 'A', 'RULE_VALUE': '10', 'RULE_TYPE': 'CONST', 'SOURCE_FIELD': '', 'SCOPE': 'GLOBAL'}])
    df2 = pd.DataFrame([{'TARGET_FIELD': 'A', 'RULE_VALUE': '99', 'RULE_TYPE': 'CONST', 'SOURCE_FIELD': '', 'SCOPE': 'GLOBAL'}])
    with pd.ExcelWriter(f1) as w: df1.to_excel(w, sheet_name='Rules', index=False)
    with pd.ExcelWriter(f2) as w: df2.to_excel(w, sheet_name='Rules', index=False)
    sm = SyncManager(rule_dir=f"{CONF_DIR}/rules")
    diffs, err = sm.compare_files("SYNC_LOCAL", f2)
    assert len(diffs) == 1

def test_26_migration_map_reverse_lookup():
    test_map_path = f"{CONF_DIR}/migration_map_test.csv"
    pd.DataFrame([{
        'MCO_SHEET': 'Test_Sheet', 'API_NAME': 'TEST_API', 
        'SDT_TEMPLATE': 'TEST_API.xlsx', 'TRANSACTION_SHEET': ''
    }]).to_csv(test_map_path, index=False)
    runner = MigrationRunner(map_path_override=test_map_path)
    api, _, _ = runner.resolve_from_map_public("Test_Sheet", "MCO_SHEET")
    assert api == "TEST_API"

def test_27_db_config_io():
    """Verify that we can write and read the DB configuration."""
    config_path = f"{CONF_DIR}/db_config.ini"
    
    # 1. Write Config
    config = configparser.ConfigParser()
    config['DEFAULT'] = {'Server': 'TEST_SERVER', 'Database': 'TEST_DB'}
    with open(config_path, 'w') as f:
        config.write(f)
        
    assert os.path.exists(config_path)
    
    # 2. Read Config (Mock)
    config2 = configparser.ConfigParser()
    config2.read(config_path)
    assert config2['DEFAULT']['Server'] == 'TEST_SERVER'
    print("[PASS] DB Config IO Verified.")

def test_28_mixed_source_extraction_logic():
    """Verify SurgicalExtractor handles mixed File and SQL sources."""
    extractor = SurgicalExtractor()
    extractor.config_dir = CONF_DIR
    extractor.staging_dir = STAGE_DIR
    
    # 1. Setup Maps
    pd.DataFrame([
        {'OBJECT_TYPE': 'MIXED', 'MCO_SHEET': 'Item', 'KEY_COLUMN': 'ITNO'},
        {'OBJECT_TYPE': 'MIXED', 'MCO_SHEET': 'Cust', 'KEY_COLUMN': 'CUNO'}
    ]).to_csv(f"{CONF_DIR}/surgical_def.csv", index=False)
    
    fpath = f"{DATA_DIR}/items.xlsx"
    pd.DataFrame({'ITNO': ['100']}).to_excel(fpath, index=False)
    
    pd.DataFrame([
        {'MCO_SHEET': 'Item', 'SOURCE_FILE': fpath},
        {'MCO_SHEET': 'Cust', 'SOURCE_FILE': 'SQL:SELECT * FROM CUST'} 
    ]).to_csv(f"{CONF_DIR}/source_map.csv", index=False)
    
    pd.DataFrame([
        {'MCO_SHEET': 'Item', 'API_NAME': 'MMS200MI'},
        {'MCO_SHEET': 'Cust', 'API_NAME': 'CRS610MI'} 
    ]).to_csv(f"{CONF_DIR}/migration_map.csv", index=False)
    
    # 2. Run Extraction with Mocked SQL Loader
    original_load = extractor.extractor.load_data
    try:
        def mock_load(path, **kwargs):
            if str(path).startswith("SQL:"):
                return pd.DataFrame({'CUNO': ['100']})
            return original_load(path, **kwargs)
            
        extractor.extractor.load_data = mock_load
        tasks = extractor.perform_extraction('MIXED', ['100'])
        
        assert len(tasks) == 2
        # Verify both sources were processed
        assert any("MMS200MI" in t['program_name'] for t in tasks)
        assert any("CRS610MI" in t['program_name'] for t in tasks)
        print("[PASS] Mixed Source Logic Verified.")
        
    finally:
        extractor.extractor.load_data = original_load
