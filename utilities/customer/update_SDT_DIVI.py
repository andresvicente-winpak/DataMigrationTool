import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog
import os
import sys

def select_file(prompt_text):
    """Opens a file dialog and returns the path."""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title=prompt_text,
        filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")]
    )
    if not file_path:
        print("No file selected. Exiting.")
        sys.exit()
    return file_path

def find_update_column(columns, suffix, drop_chars=2):
    """
    Finds a column name in a list where the name (minus first N chars) matches the suffix.
    Returns the column name if found, otherwise None.
    """
    for col in columns:
        col_str = str(col)
        if len(col_str) > drop_chars:
            trimmed_name = col_str[drop_chars:]
            if trimmed_name == suffix:
                return col_str
    return None

def main():
    print("--- Excel Updater (Format Preserving) ---\n")

    # 1. Select the SDT File
    print("Please select the SDT (Source) file...")
    sdt_path = select_file("Select SDT File (Multisheet)")
    print(f"Selected SDT: {os.path.basename(sdt_path)}")

    # 2. Select the Updates File
    print("\nPlease select the Updates file...")
    updates_path = select_file("Select Updates File")
    print(f"Selected Updates: {os.path.basename(updates_path)}")

    # ---------------------------------------------------------
    # STEP A: Read the Updates Data (using Pandas for speed)
    # ---------------------------------------------------------
    print("\nReading updates file...")
    try:
        updates_df = pd.read_excel(updates_path)
    except Exception as e:
        print(f"Error reading updates file: {e}")
        return

    # Identify Key and Value columns in Updates file
    update_key_col = find_update_column(updates_df.columns, 'CUNO', drop_chars=2)
    update_val_col = find_update_column(updates_df.columns, 'DIVI', drop_chars=2)

    if not update_key_col or not update_val_col:
        print("Error: Could not find columns ending in 'CUNO' or 'DIVI' (after dropping 2 chars) in the updates file.")
        print(f"Available columns: {list(updates_df.columns)}")
        return

    print(f"Mapping Logic: '{update_key_col}' -> '{update_val_col}'")

    # Create a clean dictionary: { String(CUNO) : New_DIVI_Value }
    # We strip whitespace and force string to ensure '100' matches 100
    update_map = dict(zip(
        updates_df[update_key_col].astype(str).str.strip(), 
        updates_df[update_val_col]
    ))

    # ---------------------------------------------------------
    # STEP B: Modify SDT File (using OpenPyXL to preserve format)
    # ---------------------------------------------------------
    print(f"Loading SDT workbook structure (this may take a moment)...")
    try:
        wb = openpyxl.load_workbook(sdt_path)
    except Exception as e:
        print(f"Error loading SDT file: {e}")
        return

    sheet_names = wb.sheetnames
    
    # User selects sheet
    print("\nAvailable Sheets:")
    for i, name in enumerate(sheet_names):
        print(f"{i + 1}: {name}")

    selected_sheet_name = None
    while selected_sheet_name is None:
        try:
            choice = int(input("\nEnter the number of the sheet to update: "))
            if 1 <= choice <= len(sheet_names):
                selected_sheet_name = sheet_names[choice - 1]
            else:
                print("Invalid number.")
        except ValueError:
            print("Please enter a valid number.")

    ws = wb[selected_sheet_name]
    print(f"Processing sheet: '{selected_sheet_name}'...")

    # Find column indices in the SDT sheet (assuming headers are in Row 1)
    # OpenPyXL columns are 1-based indices
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    
    try:
        # Convert header row to list to find index
        headers = list(header_row)
        cuno_index = headers.index('CUNO') # 0-based index
        divi_index = headers.index('DIVI') # 0-based index
    except ValueError:
        print(f"Error: Could not find exact columns 'CUNO' and 'DIVI' in row 1 of sheet '{selected_sheet_name}'.")
        return

    # Iterate rows and update
    match_count = 0
    
    # iter_rows yields tuples of cells. We start at row 2 to skip header.
    for row in ws.iter_rows(min_row=2):
        cuno_cell = row[cuno_index]
        divi_cell = row[divi_index]

        # Get CUNO value, convert to string safely
        current_cuno = str(cuno_cell.value).strip() if cuno_cell.value is not None else ""

        if current_cuno in update_map:
            # Update the DIVI cell value
            new_value = update_map[current_cuno]
            
            # Only update if value is different (optional, but good practice)
            if divi_cell.value != new_value:
                divi_cell.value = new_value
                match_count += 1

    print(f"Update complete. Updated {match_count} rows.")

    # ---------------------------------------------------------
    # STEP C: Save Result
    # ---------------------------------------------------------
    filename, ext = os.path.splitext(sdt_path)
    output_path = f"{filename}_updated{ext}"

    print(f"Saving file to: {output_path} ...")
    try:
        wb.save(output_path)
        print("Success! File saved with formatting preserved.")
    except PermissionError:
        print("Error: Permission denied. Please close the Excel file if it is open.")

if __name__ == "__main__":
    main()